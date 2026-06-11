#!/usr/bin/env python3
"""Sync Hebrew font index data from hebrew-fonts-tagging.xlsx into JSX files."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "hebrew-fonts-tagging.xlsx"
TARGETS = [
    ROOT.parent / "src/components/sections/DictionarySection.jsx",
    ROOT / "hebrew-font-index.jsx",
]

FO_MAP = {
    "אאא": "aaa",
    "פונטימונים": "fontimonim",
    "הגילדה": "hagilda",
    "פונטף": "fontef",
    "עזר טייפ האוס": "ezer",
    "הפונטיה": "hafontia",
    "ריזינגר × הפונטיה": "reisinger",
}
STYLES_MAP = {
    "סנס": "sans",
    "סריף": "serif",
    "סלאב": "slab",
    "מונו": "mono",
    "מונוספייס": "mono",
    "כתב־יד": "hand",
    "כתב־יד / ידני": "hand",
    "סטנסיל": "stencil",
}
USES_MAP = {
    "כותרות": "display",
    "כותרות / ראווה": "display",
    "טקסט רץ": "text",
    "אות ספר": "book",
    "מיתוג": "brand",
    "מיתוג ולוגו": "brand",
}
TAGS_MAP = {
    "דו־לשוני": "bilingual",
    "דו־לשוני (עב׳–אנ׳)": "bilingual",
    "תלת־לשוני": "trilingual",
    "תלת־לשוני (+ערבית)": "trilingual",
    "מולטי־לשוני": "multilingual",
    "וריאבלי": "variable",
    "ניקוד": "niqqud",
    "ניקוד מתוכנת": "niqqud",
    "חינמי": "free",
    "צר": "narrow",
    "רחב": "wide",
    "מעוגל": "rounded",
    "נוסטלגי": "nostalgic",
    "קלאסי": "classic",
    "קלאסי / מקורות": "classic",
    "עיבוד היסטורי": "revival",
    "עיבוד לאות היסטורית": "revival",
    "גראנג׳": "grunge",
}
FO_ORDER = ["aaa", "fontimonim", "hagilda", "fontef", "ezer", "hafontia", "reisinger"]


def split_vals(value) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def map_vals(values: list[str], mapping: dict[str, str]) -> list[str]:
    out: list[str] = []
    for value in values:
        key = mapping.get(value)
        if key and key not in out:
            out.append(key)
    return out


def esc(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def fmt_arr(items: list[str]) -> str:
    if not items:
        return "[]"
    return "[" + ", ".join(f'"{esc(item)}"' for item in items) + "]"


def load_existing_urls(path: Path) -> dict[tuple[str, str], dict]:
    content = path.read_text()
    existing: dict[tuple[str, str], dict] = {}
    pattern = re.compile(
        r'\{n:"([^"]+)",lat:"([^"]+)",fo:"([^"]+)",d:([^,]+),y:([^,]+),'
        r"st:\[([^\]]*)\],use:\[([^\]]*)\],tg:\[([^\]]*)\],w:([^,]+),"
        r'url:"([^"]+)",vb:\[([^\]]*)\]\}'
    )
    for match in pattern.finditer(content):
        name, _latin, foundry, _d, _y, _st, _use, _tg, weights, url, _vb = match.groups()
        existing[(name, foundry)] = {
            "w": None if weights.strip() == "null" else int(weights.strip()),
            "url": url,
        }
    return existing


def build_blocks(existing: dict[tuple[str, str], dict]) -> tuple[str, str]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    all_vibes: set[str] = set()
    fonts_by_fo: dict[str, list[dict]] = {}

    for row in range(2, ws.max_row + 1):
        foundry = FO_MAP[ws.cell(row, 2).value]
        name = ws.cell(row, 3).value
        latin = ws.cell(row, 4).value or ""
        designer = ws.cell(row, 5).value
        year = ws.cell(row, 6).value
        styles = map_vals(split_vals(ws.cell(row, 7).value), STYLES_MAP)
        uses = map_vals(split_vals(ws.cell(row, 8).value), USES_MAP)
        tags = map_vals(split_vals(ws.cell(row, 9).value), TAGS_MAP)
        vibes = split_vals(ws.cell(row, 10).value)
        all_vibes.update(vibes)

        prev = existing.get((name, foundry), {})
        fonts_by_fo.setdefault(foundry, []).append(
            {
                "n": name,
                "lat": latin or prev.get("lat", ""),
                "fo": foundry,
                "d": designer if designer else None,
                "y": int(year) if year else None,
                "st": styles,
                "use": uses,
                "tg": tags,
                "w": prev.get("w"),
                "url": prev.get("url", ""),
                "vb": vibes,
            }
        )

    vibes_lines = ["const VIBES = {"]
    for label in sorted(all_vibes):
        vibes_lines.append(f'  "{esc(label)}": "{esc(label)}",')
    vibes_lines.append("};")

    font_lines = ["const FONTS = ["]
    for foundry in FO_ORDER:
        items = fonts_by_fo[foundry]
        font_lines.append(f"  /* ---- {foundry} ({len(items)}) ---- */")
        for item in items:
            designer = "null" if item["d"] is None else f'"{esc(item["d"])}"'
            year = "null" if item["y"] is None else str(item["y"])
            weights = "null" if item["w"] is None else str(item["w"])
            font_lines.append(
                "  {"
                f'n:"{esc(item["n"])}",lat:"{esc(item["lat"])}",fo:"{item["fo"]}",'
                f"d:{designer},y:{year},"
                f'st:{fmt_arr(item["st"])},use:{fmt_arr(item["use"])},tg:{fmt_arr(item["tg"])},'
                f'w:{weights},url:"{esc(item["url"])}",vb:{fmt_arr(item["vb"])}}},'
            )
    font_lines.append("];")

    return "\n".join(vibes_lines), "\n".join(font_lines)


def apply_blocks(path: Path, vibes_block: str, fonts_block: str) -> None:
    content = path.read_text()
    content, vibes_count = re.subn(r"const VIBES = \{[\s\S]*?\};", vibes_block, content, count=1)
    content, fonts_count = re.subn(r"const FONTS = \[[\s\S]*?\];", fonts_block, content, count=1)
    if vibes_count != 1 or fonts_count != 1:
        raise RuntimeError(f"Failed to update {path}")
    path.write_text(content)


def main() -> None:
    existing = load_existing_urls(TARGETS[0])
    vibes_block, fonts_block = build_blocks(existing)
    for target in TARGETS:
        apply_blocks(target, vibes_block, fonts_block)
        print(f"Updated {target}")


if __name__ == "__main__":
    main()
